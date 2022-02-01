
import requests
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
from datetime import date, timedelta
from send_mail import send_mail
from config import API_KEY


stocks = ['AAPL', 'VWAGY', 'BNTX', 'NVDA', 'DDAIF']
stock_prices = []

try:
    if date.today().weekday() <= 4 and (date.today() - timedelta(days=1)).weekday() <= 4:
        for s in stocks:

            response = requests.get(
                'https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol=' + s + '&apikey='+API_KEY)
            data = response.json()
            stock_prices.append(float(data['Time Series (Daily)'][str(date.today()-timedelta(days=1))]['2. high']))

        wb = load_workbook('test/stockdata.xlsx')
        ws = wb.active
        row_index = 1

        while True:
            if not ws.cell(row_index, 1).value:
                break
            row_index = row_index + 1

        col_index = 2
        ws.cell(row_index, 1).value = date.today()
        for r in stock_prices:
            ws.cell(row_index, col_index).value = r
            if not isinstance(ws.cell(row_index - 1, col_index).value, str):
                if ws.cell(row_index - 1, col_index).value > r:
                    ws.cell(row_index, col_index).fill = PatternFill(patternType='solid',
                                                                     fgColor='FF0000')
                else:
                    ws.cell(row_index, col_index).fill = PatternFill(fill_type='solid',
                                                                     start_color='FF008000',
                                                                     end_color='FF008000')
            col_index = col_index + 1
        wb.save('venv/test/stockdata.xlsx')
        send_mail("Good evening Linus," + "\n\n" + "the Excel has been updated")
except Exception as e:
    print(e)
    send_mail("Good evening Linus," + "\n\n" + "this error: "+str(e)+" occured")
